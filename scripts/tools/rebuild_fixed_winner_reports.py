#!/usr/bin/env python3
"""Rebuild winner-lens reports with source-faithful R-pattern strings.

The original winner inventory is never modified. Corrected reports are written
to a separate root, with one canonical report per results-date/state/winner.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from modules.vtrac_reference import get_vtrac_index  # noqa: E402
from scripts.tools.generate_winners_from_results import parse_results  # noqa: E402
from src.core import module_c_vtrac as vtrac  # noqa: E402
from src.utils.extract_data import _clean_cell, _format_draw  # noqa: E402
from src.utils.table_generator import build_combined_table  # noqa: E402

DEFAULT_REPORTS_ROOT = ROOT / "reports" / "stable" / "winners_by_date"
DEFAULT_OUTPUT_ROOT = ROOT / "reports" / "stable" / "winners_by_date_fixed"
HISTORY_ROOT = ROOT / "data" / "history"
RESULTS_ROOT = ROOT / "data" / "results"
VARIANT_COLUMNS = {
    "Midday": ["V", "W", "X", "Y", "Z", "AA", "AB"],
    "Evening": ["AI", "AJ", "AK", "AL", "AM", "AN", "AO"],
    "Combined": ["BI", "BJ", "BK", "BL", "BM", "BN", "BO"],
}
VARIANT_INDEXES = {
    variant: [column_index_from_string(column) - 1 for column in columns]
    for variant, columns in VARIANT_COLUMNS.items()
}
SET_OFFSETS = {1: 3, 2: 49, 3: 95}
SET_SHIFTS = {1: 0, 2: 1, 3: 2}
PATTERN_ROWS = {"R2": 1, "R4": 2, "R6": 3, "R8": 4}
TABLE_COLUMNS = ["7", "6", "5", "4", "3", "2", "1"]
MAX_SOURCE_ROW = 140
MAX_SOURCE_COLUMN = max(index for indexes in VARIANT_INDEXES.values() for index in indexes) + 1
REPAIR_ID = "R_PATTERN_SOURCE_LENGTH_V1"
EXCEL_ERRORS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}


def _source_pattern(value: Any) -> str:
    text = _clean_cell(value)
    return "" if text.upper() in EXCEL_ERRORS else text


def _resolve_history_workbook(results_date: str) -> tuple[str, Path]:
    history_date = (
        datetime.strptime(results_date, "%Y-%m-%d") - timedelta(days=1)
    ).strftime("%Y-%m-%d")
    candidates = (
        HISTORY_ROOT / f"Pick3StatsC4_{history_date}.xlsm",
        HISTORY_ROOT / f"Pick3StatsC4_{history_date.replace('-', '_')}.xlsm",
    )
    workbook = next((path for path in candidates if path.exists()), None)
    if workbook is None:
        expected = ", ".join(str(path.relative_to(ROOT)) for path in candidates)
        raise FileNotFoundError(f"No history workbook for {results_date}; expected {expected}")
    return history_date, workbook


def _load_sheet_grid(workbook: Any, state: str) -> list[tuple[Any, ...]]:
    if state not in workbook.sheetnames:
        raise KeyError(f"{state} is not present in {workbook.filename}")
    worksheet = workbook[state]
    return list(
        worksheet.iter_rows(
            min_row=1,
            max_row=MAX_SOURCE_ROW,
            min_col=1,
            max_col=MAX_SOURCE_COLUMN,
            values_only=True,
        )
    )


def _extract_state_data(grid: list[tuple[Any, ...]]) -> dict[str, dict[str, Any]]:
    state_data: dict[str, dict[str, Any]] = {}
    for variant, source_indexes in VARIANT_INDEXES.items():
        section: dict[str, Any] = {"Set1": {}, "Set2": {}, "Set3": {}}
        for set_number in (1, 2, 3):
            set_name = f"Set{set_number}"
            draw_numbers: Iterable[int] = range(1, 8) if set_number == 1 else (1,)
            for draw_number in draw_numbers:
                base = SET_OFFSETS[set_number] - SET_SHIFTS[set_number]
                if set_number == 1:
                    base += (draw_number - 1) * 6
                    active_indexes = source_indexes[draw_number - 1 :]
                else:
                    active_indexes = source_indexes
                draw_payload: dict[str, list[str]] = {
                    "draw_data": [_format_draw(grid[base][index]) for index in active_indexes]
                }
                for row_type, row_offset in PATTERN_ROWS.items():
                    draw_payload[row_type] = [
                        _source_pattern(grid[base + row_offset][index])
                        for index in active_indexes
                    ]
                section[set_name][f"Draw{draw_number}"] = draw_payload
        state_data[variant] = section
    return state_data


def _build_tables(grid: list[tuple[Any, ...]]) -> dict[str, Any]:
    state_data = _extract_state_data(grid)
    return {
        f"{variant}_combined": build_combined_table(state_data[variant])
        for variant in ("Midday", "Evening", "Combined")
    }


def _strip_markers(value: Any) -> str:
    text = str(value or "").strip().rstrip("*")
    return "" if text == "N/A" else text


def _table_integrity(tables: dict[str, Any]) -> dict[str, int]:
    counts: Counter[str] = Counter()
    for table in tables.values():
        for _, row in table.iterrows():
            row_type = str(row.get("RowType") or "")
            for column in TABLE_COLUMNS:
                value = _strip_markers(row.get(column))
                if value.lower() in {"nan", "<na>", "none"}:
                    raise ValueError(f"Literal missing-value token found in {row_type} C{column}")
                if row_type in PATTERN_ROWS and value:
                    counts["pattern_cells"] += 1
                    if value.isdigit() and len(value) < 3:
                        counts["short_pattern_cells"] += 1

        pattern_table = table[table["RowType"].isin(PATTERN_ROWS)]
        for (_, _), group in pattern_table.groupby(["Set", "Draw"], sort=False):
            for column in ("1", "2"):
                values_by_row = {
                    str(row["RowType"]): _strip_markers(row[column])
                    for _, row in group.iterrows()
                }
                if set(values_by_row) != set(PATTERN_ROWS):
                    continue
                values = [values_by_row[row_type] for row_type in PATTERN_ROWS]
                if values[0] and len(set(values)) == 1:
                    counts["strict_end_box_consensus"] += 1
                    if len(values[0]) == 1:
                        counts["strict_single_digit_consensus"] += 1
    return dict(counts)


def _source_state_dirs(reports_root: Path, results_date: str) -> list[str]:
    date_root = reports_root / results_date
    if not date_root.exists():
        return []
    return sorted(path.name for path in date_root.iterdir() if path.is_dir())


def _selected_dates(reports_root: Path, requested: list[str] | None) -> list[str]:
    if requested:
        return sorted(dict.fromkeys(requested))
    return sorted(
        path.name
        for path in reports_root.iterdir()
        if path.is_dir() and re.fullmatch(r"\d{4}-\d{2}-\d{2}", path.name)
    )


def _write_report(
    *,
    state: str,
    winner: str,
    tables: dict[str, Any],
    output_dir: Path,
    results_date: str,
    history_date: str,
    workbook_path: Path,
    force: bool,
) -> tuple[Path, Path, bool]:
    index = get_vtrac_index(winner)
    patterns = set(vtrac.get_all_combinations_for_index(index))
    stem = f"{state}_vtrac{index}_winner_{winner}_FIXED_R_PATTERN_V1"
    html_path = output_dir / f"{stem}.html"
    json_path = output_dir / f"{stem}.json"
    if not force and html_path.exists() and json_path.exists():
        return html_path, json_path, False

    html = vtrac.generate_index_html_report(
        state,
        index,
        patterns,
        tables,
        score=0,
        rank=0,
        timestamp=REPAIR_ID,
        winner_combo=winner,
    )
    payload = vtrac.generate_index_json_report(
        state,
        index,
        patterns,
        tables,
        score=0,
        rank=0,
        timestamp=REPAIR_ID,
        winner_combo=winner,
    )
    payload["repair"] = {
        "repair_id": REPAIR_ID,
        "results_date": results_date,
        "history_date": history_date,
        "source_workbook": str(workbook_path.relative_to(ROOT)),
        "source_values_preserved": True,
        "original_reports_overwritten": False,
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    html_path.write_text(html, encoding="utf-8")
    json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return html_path, json_path, True


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reports-root", type=Path, default=DEFAULT_REPORTS_ROOT)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--date", action="append", dest="dates")
    parser.add_argument("--state", action="append", dest="states")
    parser.add_argument(
        "--winner",
        action="append",
        dest="winners",
        help="Override winners; valid only with exactly one --date and one --state.",
    )
    parser.add_argument("--force", action="store_true", help="Overwrite matching fixed files.")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    reports_root = args.reports_root.resolve()
    output_root = args.output_root.resolve()
    if reports_root == output_root:
        raise SystemExit("Output root must differ from the original reports root.")
    if args.winners and (not args.dates or len(args.dates) != 1 or not args.states or len(args.states) != 1):
        raise SystemExit("--winner requires exactly one --date and one --state.")

    records: list[dict[str, Any]] = []
    summary: Counter[str] = Counter()
    for results_date in _selected_dates(reports_root, args.dates):
        source_states = _source_state_dirs(reports_root, results_date)
        selected_states = list(args.states) if args.states else source_states
        if not selected_states:
            records.append({"date": results_date, "status": "empty_source_date"})
            summary["empty_source_dates"] += 1
            continue
        try:
            history_date, workbook_path = _resolve_history_workbook(results_date)
        except FileNotFoundError as exc:
            records.append(
                {
                    "date": results_date,
                    "status": "missing_history_workbook",
                    "error": str(exc),
                }
            )
            summary["missing_history_dates"] += 1
            print(f"[SKIP] {exc}")
            continue
        results_path = RESULTS_ROOT / f"{results_date}.txt"
        if not results_path.exists():
            records.append({"date": results_date, "status": "missing_results"})
            summary["missing_results_dates"] += 1
            continue
        winners_by_state = parse_results(results_path)
        if args.dry_run:
            print(
                f"[DRY] D={results_date} H={history_date} "
                f"states={len(selected_states)} workbook={workbook_path.name}"
            )
            summary["dry_run_dates"] += 1
            continue

        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
            keep_vba=True,
        )
        try:
            for state in selected_states:
                winners = list(args.winners) if args.winners else winners_by_state.get(state, [])
                winners = [winner for winner in dict.fromkeys(winners) if re.fullmatch(r"\d{3}", winner)]
                if not winners:
                    records.append(
                        {"date": results_date, "state": state, "status": "no_valid_result"}
                    )
                    summary["states_without_results"] += 1
                    continue
                try:
                    grid = _load_sheet_grid(workbook, state)
                    tables = _build_tables(grid)
                    integrity = _table_integrity(tables)
                except Exception as exc:
                    records.append(
                        {
                            "date": results_date,
                            "state": state,
                            "status": "table_build_failed",
                            "error": str(exc),
                        }
                    )
                    summary["table_build_failures"] += 1
                    continue

                state_output = output_root / results_date / state
                written = 0
                outputs = []
                for winner in winners:
                    try:
                        html_path, json_path, created = _write_report(
                            state=state,
                            winner=winner,
                            tables=tables,
                            output_dir=state_output,
                            results_date=results_date,
                            history_date=history_date,
                            workbook_path=workbook_path,
                            force=args.force,
                        )
                    except Exception as exc:
                        outputs.append({"winner": winner, "status": "failed", "error": str(exc)})
                        summary["report_failures"] += 1
                        continue
                    written += int(created)
                    summary["reports_created"] += int(created)
                    summary["reports_reused"] += int(not created)
                    outputs.append(
                        {
                            "winner": winner,
                            "status": "created" if created else "existing",
                            "html": str(html_path.relative_to(ROOT)),
                            "json": str(json_path.relative_to(ROOT)),
                        }
                    )
                summary["states_completed"] += 1
                records.append(
                    {
                        "date": results_date,
                        "history_date": history_date,
                        "state": state,
                        "status": "complete",
                        "reports_written": written,
                        "integrity": integrity,
                        "outputs": outputs,
                    }
                )
        finally:
            workbook.close()

    if not args.dry_run:
        output_root.mkdir(parents=True, exist_ok=True)
        manifest = {
            "repair_id": REPAIR_ID,
            "generated_at": datetime.now().astimezone().isoformat(),
            "original_reports_root": str(reports_root.relative_to(ROOT)),
            "fixed_reports_root": str(output_root.relative_to(ROOT)),
            "original_reports_overwritten": False,
            "summary": dict(summary),
            "records": records,
        }
        (output_root / "REBUILD_MANIFEST.json").write_text(
            json.dumps(manifest, indent=2),
            encoding="utf-8",
        )
        (output_root / "README.md").write_text(
            "# Fixed Winner Reports\n\n"
            f"Repair: `{REPAIR_ID}`\n\n"
            "These reports were rebuilt from the prior-day historical workbooks. "
            "Actual draws remain three digits; R2/R4/R6/R8 reduction strings retain "
            "their source length. The original `winners_by_date` inventory was not modified.\n",
            encoding="utf-8",
        )
    print(json.dumps(dict(summary), sort_keys=True))
    return 0 if not summary["table_build_failures"] and not summary["report_failures"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
