"""
excel_export.py - Utility functions for exporting formatted Excel files
"""

import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def export_state_tables(state_name, midday_df, evening_df, combined_df, output_dir):
    """
    Export state tables to a formatted Excel file with side-by-side layout
    
    Args:
        state_name (str): Name of the state
        midday_df (pd.DataFrame): Midday table
        evening_df (pd.DataFrame): Evening table
        combined_df (pd.DataFrame): Combined table
        output_dir (str): Directory to save the Excel file
    """
    # Create timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{state_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Calculate starting columns for each section
    col_width = len(midday_df.columns) if midday_df is not None else 0
    spacing = 2  # Number of columns to leave blank between tables
    
    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Write each table with proper spacing
        if midday_df is not None:
            midday_df.to_excel(writer, sheet_name='State Tables', startcol=0, index=False)
        
        if evening_df is not None:
            evening_df.to_excel(writer, sheet_name='State Tables', startcol=col_width + spacing, index=False)
        
        if combined_df is not None:
            combined_df.to_excel(writer, sheet_name='State Tables', startcol=(col_width + spacing) * 2, index=False)
        
        # Get the workbook and active sheet
        workbook = writer.book
        worksheet = writer.sheets['State Tables']
        
        # Add headers for each section
        header_font = Font(bold=True, size=14)
        worksheet.cell(row=1, column=1, value="Midday").font = header_font
        worksheet.cell(row=1, column=col_width + spacing + 1, value="Evening").font = header_font
        worksheet.cell(row=1, column=(col_width + spacing) * 2 + 1, value="Combined").font = header_font
        
        # Move the table data down one row to accommodate headers
        for row in range(worksheet.max_row, 1, -1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                worksheet.cell(row=row + 1, column=col, value=cell.value)
                if cell.has_style:
                    worksheet.cell(row=row + 1, column=col)._style = cell._style
        
        # Apply styling
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders and header styling
        for col in range(1, worksheet.max_column + 1):
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if row == 2:  # Header row
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
        
        # Set column widths
        for col in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    return filepath

def setup_logging_directories():
    """
    Set up necessary directories for logging if they don't exist
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    archive_dir = os.path.join(base_dir, 'data', 'archive')
    os.makedirs(archive_dir, exist_ok=True)
    return archive_dir 